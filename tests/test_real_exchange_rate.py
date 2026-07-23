from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from argentina_economic_data.inflation import Artifact, PipelineError
from argentina_economic_data.real_exchange_rate import BASE_LABEL, SERIES, SHEET_NAME, extract


def _artifact(tmp_path: Path, *, missing: str | None = None) -> Artifact:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.cell(1, 1, BASE_LABEL)
    headers = ["Período", *[header for header in SERIES if header != missing]]
    for column, header in enumerate(headers, start=1):
        sheet.cell(2, column, header)
    for row_index, period in enumerate((datetime(2026, 7, 22), datetime(2026, 7, 23)), start=3):
        sheet.cell(row_index, 1, period)
        for column in range(2, len(headers) + 1):
            sheet.cell(row_index, column, 100 + row_index + column / 10)
    path = tmp_path / "ITCRMSerie.xlsx"
    workbook.save(path)
    return Artifact("test", "https://example.test/ITCRMSerie.xlsx", path, "hash", path.stat().st_size, "now")


def test_extracts_itcrm_and_four_bilateral_series(tmp_path: Path):
    records = extract(_artifact(tmp_path))
    assert len(records) == 10
    assert {row["series_id"] for row in records} == set(SERIES.values())
    assert {row["period"] for row in records} == {"2026-07-22", "2026-07-23"}
    assert all(row["frequency"] == "daily" for row in records)
    assert all(row["unit"] == "index_dec_17_2015_100" for row in records)
    assert all(row["status"] == "official_provisional" for row in records)


def test_rejects_missing_bilateral_column(tmp_path: Path):
    with pytest.raises(PipelineError, match="faltan columnas"):
        extract(_artifact(tmp_path, missing="ITCRB China"))
