import unittest
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from argentina_economic_data.credit import (
    _monthly_gdp_denominators,
    _official_private_credit_gdp,
    _period,
    _private_currency_values,
)
from argentina_economic_data.inflation import Artifact


class CreditTests(unittest.TestCase):
    def test_bcra_decimal_period(self):
        self.assertEqual(_period(2026.05), "2026-05")
        self.assertIsNone(_period(2026.13))

    def test_gdp_denominator_interpolates_annualized_quarters(self):
        gdp = {
            "2024-Q1": Decimal("100"),
            "2024-Q2": Decimal("160"),
        }
        denominators = _monthly_gdp_denominators(gdp)
        self.assertEqual(denominators[2024 * 12 + 1], Decimal("100"))
        self.assertEqual(denominators[2024 * 12 + 2], Decimal("120"))
        self.assertEqual(denominators[2024 * 12 + 3], Decimal("140"))
        self.assertEqual(denominators[2024 * 12 + 4], Decimal("160"))

    def test_official_ratio_anchor_replaces_monthly_denominator(self):
        denominators = _monthly_gdp_denominators(
            {"2024-Q1": Decimal("100"), "2024-Q2": Decimal("160")},
            {"2024-03": Decimal("130")},
        )
        self.assertEqual(denominators[2024 * 12 + 2], Decimal("130"))

    def test_official_private_credit_ratio_adds_peso_and_dollar_loans(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "indicators.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Principales Variables"
            sheet.cell(4, 13).value = __import__("datetime").datetime(2026, 7, 31)
            sheet.cell(29, 1).value = "Préstamos al sector privado no financiero"
            sheet.cell(29, 13).value = 0.09237792973546574
            sheet.cell(45, 1).value = "Préstamos al sector privado no financiero en dólares"
            sheet.cell(45, 13).value = 0.032440397134843195
            workbook.save(path)
            artifact = Artifact("test", "https://example.test/source.xlsx", path, "abc", path.stat().st_size, "2026-08-20T00:00:00Z")

            records = _official_private_credit_gdp(artifact)

            by_series = {record["series_id"]: record for record in records}
            self.assertEqual(len(by_series), 3)
            self.assertEqual(by_series["bcra_private_nonfinancial_credit_ars_gdp_ratio"]["value"], "9.237793")
            self.assertEqual(by_series["bcra_private_nonfinancial_credit_fx_ars_gdp_ratio"]["value"], "3.244040")
            self.assertEqual(by_series["bcra_private_nonfinancial_credit_gdp_ratio"]["value"], "12.481833")
            self.assertTrue(all(record["period"] == "2026-07" for record in records))
            self.assertTrue(all(record["status"] == "official" for record in records))

    def test_private_credit_currency_opening_sums_to_total(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "series.xlsm"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "PRESTAMOS"
            row = [None] * 22
            row[0] = __import__("datetime").datetime(2026, 7, 31)
            row[8] = 100
            row[16] = 20
            row[18] = 30
            row[20] = 130
            row[21] = "PM"
            sheet.append([None] * 22)
            for _ in range(8):
                sheet.append([None] * 22)
            sheet.append(row)
            workbook.save(path)
            artifact = Artifact("test", "https://example.test/series.xlsm", path, "abc", path.stat().st_size, "2026-08-20T00:00:00Z")

            values = _private_currency_values(artifact)

            self.assertEqual(values["2026-07"]["ars"], Decimal("100"))
            self.assertEqual(values["2026-07"]["fx_usd"], Decimal("20"))
            self.assertEqual(values["2026-07"]["fx_ars"], Decimal("30"))
            self.assertEqual(values["2026-07"]["total"], Decimal("130"))


if __name__ == "__main__":
    unittest.main()
