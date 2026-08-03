from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from argentina_economic_data.inflation import Artifact
from argentina_economic_data.profit_dividends import (
    ANNUAL_SERIES,
    MONTHLY_SERIES,
    SERIES_CODE,
    SERIES_HEADER,
    SHEET_NAME,
    aggregate_annual,
    extract,
)


def _artifact(tmp_path: Path) -> Artifact:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.cell(9, 4, SERIES_HEADER)
    sheet.cell(11, 4, SERIES_CODE)
    sheet.cell(12, 1, datetime(2025, 1, 1))
    sheet.cell(12, 4, -87.5)
    sheet.cell(13, 1, datetime(2025, 2, 1))
    sheet.cell(13, 4, 0.01)
    # Los totales anuales incrustados en la fuente no deben interpretarse como meses.
    sheet.cell(14, 1, 2025)
    sheet.cell(14, 4, -87.49)
    path = tmp_path / "balance-cambiario.xlsx"
    workbook.save(path)
    return Artifact("test", "https://example.test/source.xlsx", path, "hash", path.stat().st_size, "now")


def test_extracts_monthly_outflows_and_normalizes_sign(tmp_path: Path):
    records = extract(_artifact(tmp_path))
    assert [(row["period"], Decimal(row["value"])) for row in records] == [
        ("2025-01", Decimal("87.500000")),
        ("2025-02", Decimal("-0.010000")),
    ]
    assert {row["series_id"] for row in records} == {MONTHLY_SERIES}
    assert all(row["frequency"] == "monthly" for row in records)


def test_annual_totals_require_twelve_calendar_months():
    template = {
        "series_id": MONTHLY_SERIES,
        "frequency": "monthly",
        "value": "10",
        "unit": "million_usd",
        "status": "official",
        "source_id": "x",
        "source_url": "x",
        "source_sha256": "x",
        "retrieved_at": "x",
    }
    monthly = [template | {"period": f"2024-{month:02d}"} for month in range(1, 13)]
    monthly += [template | {"period": f"2025-{month:02d}"} for month in range(1, 7)]
    annual = [row for row in aggregate_annual(monthly) if row["series_id"] == ANNUAL_SERIES]
    assert [(row["period"], Decimal(row["value"])) for row in annual] == [
        ("2024", Decimal("120.000000")),
    ]
    assert annual[0]["status"] == "calculated"
