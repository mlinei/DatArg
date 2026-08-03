from __future__ import annotations

import csv
import json
import os
import tempfile
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .inflation import OUTPUT_COLUMNS, Artifact, PipelineError, acquire

SOURCE_URL = (
    "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/"
    "informes/anexo-estadistico-mercado-cambios-balance-cambiario.xlsx"
)
LANDING_URL = "https://www.bcra.gob.ar/estadisticas-estandarizadas-sobre-la-evolucion-del-mercado-de-cambios/"
SOURCE_ID = "bcra_fx_market_profit_dividends"
SHEET_NAME = "Mercado de Cambios"
SERIES_CODE = "mlc057"
SERIES_HEADER = "Utilidades y Dividendos"
MONTHLY_SERIES = "bcra_profit_dividend_outflows_monthly"
ANNUAL_SERIES = "bcra_profit_dividend_outflows_annual"
FIRST_PERIOD = "2003-01"


def _text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _period(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    # La hoja agrega totales anuales como enteros al final. No son observaciones mensuales.
    return None


def extract(artifact: Artifact) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(artifact.path, read_only=True, data_only=True)
    except Exception as exc:
        raise PipelineError(f"utilidades y dividendos BCRA: no se pudo abrir el XLSX: {exc}") from exc
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise PipelineError(f"utilidades y dividendos BCRA: no existe la hoja {SHEET_NAME!r}")

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header_rows: dict[int, tuple[object, ...]] = {}
    code_columns: list[int] = []
    records: list[dict[str, str]] = []
    seen_periods: set[str] = set()

    try:
        for row_number, row in enumerate(rows, start=1):
            if row_number <= 11:
                header_rows[row_number] = row
            if row_number == 11:
                code_columns = [
                    index for index, value in enumerate(row)
                    if _text(value) == SERIES_CODE
                ]
                if len(code_columns) != 1:
                    raise PipelineError(
                        f"utilidades y dividendos BCRA: se esperaba una columna {SERIES_CODE}, "
                        f"se encontraron {len(code_columns)}"
                    )
                column = code_columns[0]
                if column >= len(header_rows[9]) or _text(header_rows[9][column]) != SERIES_HEADER:
                    raise PipelineError("utilidades y dividendos BCRA: cambió el encabezado de la serie")
                continue
            if row_number < 12:
                continue

            period = _period(row[0] if row else None)
            if period is None:
                continue
            if period in seen_periods:
                raise PipelineError(f"utilidades y dividendos BCRA: período duplicado: {period}")
            seen_periods.add(period)
            raw_value = row[code_columns[0]] if code_columns[0] < len(row) else None
            try:
                # En la hoja de Mercado de Cambios, los egresos se publican con signo negativo.
                # Se invierte el signo para que un giro al exterior se visualice como flujo positivo.
                value = -Decimal(str(raw_value))
            except (InvalidOperation, TypeError) as exc:
                raise PipelineError(
                    f"utilidades y dividendos BCRA: valor inválido en {period}: {raw_value!r}"
                ) from exc
            records.append({
                "series_id": MONTHLY_SERIES,
                "period": period,
                "frequency": "monthly",
                "value": format(value.quantize(Decimal("0.000001")), "f"),
                "unit": "million_usd",
                "status": "official",
                "source_id": SOURCE_ID,
                "source_url": artifact.url,
                "source_sha256": artifact.sha256,
                "retrieved_at": artifact.retrieved_at,
            })
    finally:
        workbook.close()

    if not records:
        raise PipelineError("utilidades y dividendos BCRA: la fuente no contiene observaciones mensuales")
    return records


def aggregate_annual(monthly: list[dict[str, str]]) -> list[dict[str, str]]:
    by_year: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in monthly:
        by_year[row["period"][:4]].append(row)

    records = list(monthly)
    for year, rows in sorted(by_year.items()):
        months = {row["period"][5:7] for row in rows}
        if months != {f"{month:02d}" for month in range(1, 13)}:
            continue
        value = sum((Decimal(row["value"]) for row in rows), Decimal())
        records.append(rows[-1] | {
            "series_id": ANNUAL_SERIES,
            "period": year,
            "frequency": "annual",
            "value": format(value.quantize(Decimal("0.000001")), "f"),
            "status": "calculated",
        })
    return records


def promote(records: list[dict[str, str]], root: Path, run_id: str) -> dict[str, object]:
    records.sort(key=lambda row: (row["series_id"], row["period"]))
    keys = [(row["series_id"], row["period"]) for row in records]
    if len(keys) != len(set(keys)):
        raise PipelineError("utilidades y dividendos BCRA: claves duplicadas")

    monthly = [row for row in records if row["series_id"] == MONTHLY_SERIES]
    annual = [row for row in records if row["series_id"] == ANNUAL_SERIES]
    if monthly[0]["period"] != FIRST_PERIOD or len(monthly) < 250:
        raise PipelineError("utilidades y dividendos BCRA: cobertura histórica inesperada")
    periods = [row["period"] for row in monthly]
    expected_periods: list[str] = []
    year, month = map(int, FIRST_PERIOD.split("-"))
    while f"{year:04d}-{month:02d}" <= periods[-1]:
        expected_periods.append(f"{year:04d}-{month:02d}")
        month += 1
        if month == 13:
            year += 1
            month = 1
    if periods != expected_periods:
        raise PipelineError("utilidades y dividendos BCRA: la cobertura mensual tiene faltantes")
    if not annual:
        raise PipelineError("utilidades y dividendos BCRA: faltan los totales de años completos")

    target_dir = root / "data" / "processed"
    log_dir = root / "data" / "logs" / "profit_dividends"
    target_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / "profit_dividends.csv"

    old: dict[tuple[str, str], str] = {}
    if target.exists():
        with target.open(encoding="utf-8", newline="") as handle:
            old = {(row["series_id"], row["period"]): row["value"] for row in csv.DictReader(handle)}
    new = {(row["series_id"], row["period"]): row["value"] for row in records}
    report = {
        "run_id": run_id,
        "rows": len(records),
        "series": len({row["series_id"] for row in records}),
        "min_period": monthly[0]["period"],
        "max_period": monthly[-1]["period"],
        "created": len(new.keys() - old.keys()),
        "deleted": len(old.keys() - new.keys()),
        "modified": sum(old[key] != new[key] for key in old.keys() & new.keys()),
    }
    if old.keys() - new.keys():
        raise PipelineError("utilidades y dividendos BCRA: la fuente eliminó observaciones existentes")

    fd, temporary = tempfile.mkstemp(prefix="profit-dividends-", suffix=".csv", dir=target_dir)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
            writer.writeheader()
            writer.writerows(records)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)

    (log_dir / f"{run_id}.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return report


def run(root: Path, source_file: Path | None = None) -> dict[str, object]:
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    artifact = acquire(SOURCE_ID, SOURCE_URL, root / "data" / "raw", source_file)
    return promote(aggregate_annual(extract(artifact)), root, run_id)
