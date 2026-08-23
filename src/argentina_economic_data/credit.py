from __future__ import annotations

import csv
import hashlib
import json
import os
import tempfile
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from .inflation import OUTPUT_COLUMNS, Artifact, PipelineError, acquire

SOURCE_PAGE = "https://www.bcra.gob.ar/prestamos-y-otros-activos-de-las-entidades-financieras/"
PRIVATE_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/perser_priv.xls"
PRIVATE_CURRENCY_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/series.xlsm"
PUBLIC_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/perser_pub.xls"
SECURITIES_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/titpubser.xls"
MONTHLY_INDICATORS_URL = (
    "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/"
    "indicadores-informe-monetario-mensual-{period}.xlsx"
)
REAL_BASE_PERIOD = "2019-12"
LEVEL_SERIES = {
    "bcra_private_nonfinancial_credit",
    "bcra_private_nonfinancial_credit_ars",
    "bcra_private_nonfinancial_credit_fx_ars",
    "bcra_public_loans_government",
    "bcra_public_loans_enterprises",
    "bcra_public_loans_total",
    "bcra_public_exposure_securities",
    "bcra_public_exposure_total",
}


def _private_currency_values(artifact: Artifact) -> dict[str, dict[str, Decimal]]:
    """Promedios mensuales de préstamos privados por moneda del I.M.D. del BCRA."""
    try:
        workbook = load_workbook(artifact.path, data_only=True, read_only=True)
        sheet = workbook["PRESTAMOS"]
    except Exception as exc:
        raise PipelineError(f"crédito BCRA: no se pudo leer la apertura por moneda: {exc}") from exc

    result: dict[str, dict[str, Decimal]] = {}
    for row in sheet.iter_rows(min_row=10, values_only=True):
        if len(row) <= 21 or row[21] != "PM" or not isinstance(row[0], datetime):
            continue
        period = row[0].strftime("%Y-%m")
        try:
            result[period] = {
                "ars": Decimal(str(row[8])),
                "fx_usd": Decimal(str(row[16])),
                "fx_ars": Decimal(str(row[18])),
                "total": Decimal(str(row[20])),
            }
        except (InvalidOperation, TypeError) as exc:
            raise PipelineError(f"crédito BCRA: apertura por moneda inválida en {period}") from exc
    if not result:
        raise PipelineError("crédito BCRA: no hay promedios mensuales en la apertura por moneda")
    for period, values in result.items():
        if abs(values["ars"] + values["fx_ars"] - values["total"]) > Decimal("0.1"):
            raise PipelineError(f"crédito BCRA: los componentes no suman el total en {period}")
    return result


def _period(value: object) -> str | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    year = int(number)
    month = round((number - year) * 100)
    return f"{year:04d}-{month:02d}" if 1 <= month <= 12 else None


def _number(value: object, field: str, period: str) -> Decimal:
    try:
        number = Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise PipelineError(f"crédito BCRA: valor inválido en {field}/{period}") from exc
    if number < 0:
        raise PipelineError(f"crédito BCRA: valor negativo en {field}/{period}")
    return number / Decimal(1000)  # miles de pesos -> millones de pesos


def _sheet(artifact: Artifact, name: str):
    try:
        workbook = xlrd.open_workbook(artifact.path)
        return workbook.sheet_by_name(name)
    except Exception as exc:
        raise PipelineError(f"crédito BCRA: no se pudo leer la hoja {name}: {exc}") from exc


def _values(artifact: Artifact, sheet_name: str, columns: dict[str, tuple[int, ...]]) -> dict[str, dict[str, Decimal]]:
    sheet = _sheet(artifact, sheet_name)
    result: dict[str, dict[str, Decimal]] = {}
    for row in range(26, sheet.nrows):
        period = _period(sheet.cell_value(row, 0))
        if not period or not str(sheet.cell_value(row, 1)).strip():
            continue
        values = {
            name: sum((_number(sheet.cell_value(row, column), name, period) for column in indices), Decimal())
            for name, indices in columns.items()
        }
        result[period] = values
    if not result:
        raise PipelineError(f"crédito BCRA: la hoja {sheet_name} no contiene observaciones")
    return result


def _record(series_id: str, period: str, value: Decimal, status: str, artifact: Artifact, source_hash: str) -> dict[str, str]:
    return {
        "series_id": series_id, "period": period, "frequency": "monthly",
        "value": format(value.quantize(Decimal("0.000001")), "f"), "unit": "million_ars",
        "status": status, "source_id": "bcra_financial_system_credit",
        "source_url": SOURCE_PAGE, "source_sha256": source_hash,
        "retrieved_at": artifact.retrieved_at,
    }


def extract(private: Artifact, private_currency: Artifact, public: Artifact, securities: Artifact) -> list[dict[str, str]]:
    private_values = _values(private, "Prest-Tot", {"private": (3,)})
    currency_values = _private_currency_values(private_currency)
    public_values = _values(public, "Series", {
        "government": (3, 27),
        "public_enterprises": (10, 11, 12, 13, 34, 35, 36, 37),
    })
    security_values = _values(securities, "Datos", {
        # Gobierno nacional + gobiernos provinciales y municipales. Se excluyen títulos del BCRA.
        "government_securities": (9, 12),
    })
    source_hash = hashlib.sha256(
        "|".join((private.sha256, public.sha256, securities.sha256)).encode()
    ).hexdigest()
    records: list[dict[str, str]] = []
    for period, values in private_values.items():
        if period in currency_values:
            continue
        records.append(_record("bcra_private_nonfinancial_credit", period, values["private"], "official", private, source_hash))
    for period, values in currency_values.items():
        common = {
            "period": period, "frequency": "monthly", "status": "official",
            "source_id": "bcra_daily_monetary_report_monthly_average",
            "source_url": private_currency.url, "source_sha256": private_currency.sha256,
            "retrieved_at": private_currency.retrieved_at,
        }
        records.extend([
            common | {"series_id": "bcra_private_nonfinancial_credit_ars", "value": format(values["ars"], "f"), "unit": "million_ars"},
            common | {"series_id": "bcra_private_nonfinancial_credit_fx_usd", "value": format(values["fx_usd"], "f"), "unit": "million_usd"},
            common | {"series_id": "bcra_private_nonfinancial_credit_fx_ars", "value": format(values["fx_ars"], "f"), "unit": "million_ars"},
            common | {"series_id": "bcra_private_nonfinancial_credit", "value": format(values["total"], "f"), "unit": "million_ars"},
        ])
    for period, values in public_values.items():
        government = values["government"]
        enterprises = values["public_enterprises"]
        total = government + enterprises
        records.extend([
            _record("bcra_public_loans_government", period, government, "official", public, source_hash),
            _record("bcra_public_loans_enterprises", period, enterprises, "official", public, source_hash),
            _record("bcra_public_loans_total", period, total, "calculated", public, source_hash),
        ])
    common_periods = sorted(set(public_values) & set(security_values))
    for period in common_periods:
        loans = public_values[period]["government"] + public_values[period]["public_enterprises"]
        securities_value = security_values[period]["government_securities"]
        records.extend([
            _record("bcra_public_exposure_securities", period, securities_value, "official", securities, source_hash),
            _record("bcra_public_exposure_total", period, loans + securities_value, "calculated", securities, source_hash),
        ])

    # Variaciones nominales interanuales. Se publican explícitamente como nominales.
    levels = {(row["series_id"], row["period"]): Decimal(row["value"]) for row in records}
    for row in list(records):
        previous_period = f"{int(row['period'][:4]) - 1:04d}{row['period'][4:]}"
        previous = levels.get((row["series_id"], previous_period))
        current = Decimal(row["value"])
        if previous and previous > 0:
            records.append(row | {
                "series_id": f"{row['series_id']}_nominal_yoy",
                "value": format(((current / previous - 1) * 100).quantize(Decimal("0.000001")), "f"),
                "unit": "percent_change",
                "status": "calculated",
            })
    return records


def _official_private_credit_gdp(artifact: Artifact) -> list[dict[str, str]]:
    """Extrae el crédito privado total/PIB publicado en el Informe Monetario.

    El cuadro del BCRA informa por separado préstamos en pesos y en moneda
    extranjera. Ambas razones ya usan el PIB mensual estimado por el BCRA, por
    lo que se suman sin convertir saldos ni aplicar un denominador de DatArg.
    """
    try:
        workbook = load_workbook(artifact.path, data_only=True, read_only=True)
        sheet = workbook["Principales Variables"]
    except Exception as exc:
        raise PipelineError(f"crédito BCRA: no se pudo leer el informe monetario: {exc}") from exc

    wanted = {
        "Préstamos al sector privado no financiero": None,
        "Préstamos al sector privado no financiero en dólares": None,
    }
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "").strip()
        if label in wanted:
            wanted[label] = row
    missing = [label for label, row in wanted.items() if row is None]
    if missing:
        raise PipelineError(f"crédito BCRA: faltan filas del informe monetario: {', '.join(missing)}")

    records: list[dict[str, str]] = []
    for column in range(13, sheet.max_column + 1):
        date_value = sheet.cell(4, column).value
        if not isinstance(date_value, datetime):
            continue
        period = date_value.strftime("%Y-%m")
        values: list[Decimal] = []
        for row in wanted.values():
            raw = sheet.cell(row, column).value  # type: ignore[arg-type]
            try:
                values.append(Decimal(str(raw)))
            except (InvalidOperation, TypeError) as exc:
                raise PipelineError(f"crédito BCRA: razón crédito/PIB inválida en {period}") from exc
        ratios = {
            "bcra_private_nonfinancial_credit_ars_gdp_ratio": values[0] * Decimal(100),
            "bcra_private_nonfinancial_credit_fx_ars_gdp_ratio": values[1] * Decimal(100),
            "bcra_private_nonfinancial_credit_gdp_ratio": sum(values, Decimal()) * Decimal(100),
        }
        for series_id, ratio in ratios.items():
            records.append({
            "series_id": series_id,
            "period": period,
            "frequency": "monthly",
            "value": format(ratio.quantize(Decimal("0.000001")), "f"),
            "unit": "percent_gdp",
            "status": "official",
            "source_id": "bcra_monthly_monetary_indicators",
            "source_url": artifact.url,
            "source_sha256": artifact.sha256,
            "retrieved_at": artifact.retrieved_at,
            })
    if not records:
        raise PipelineError("crédito BCRA: el informe monetario no contiene razones crédito/PIB")
    return records


def _previous_months(count: int = 3) -> list[str]:
    now = datetime.now(timezone.utc)
    year, month = now.year, now.month
    periods: list[str] = []
    for _ in range(count):
        month -= 1
        if month == 0:
            year -= 1
            month = 12
        periods.append(f"{year:04d}-{month:02d}")
    return periods


def _acquire_monthly_indicators(raw: Path, local: Path | None) -> Artifact:
    if local:
        period = _previous_months(1)[0]
        return acquire(
            "bcra_monthly_monetary_indicators_xlsx",
            MONTHLY_INDICATORS_URL.format(period=period), raw, local,
        )
    errors: list[str] = []
    for period in _previous_months():
        url = MONTHLY_INDICATORS_URL.format(period=period)
        try:
            return acquire("bcra_monthly_monetary_indicators_xlsx", url, raw)
        except PipelineError as exc:
            errors.append(f"{period}: {exc}")
    raise PipelineError("crédito BCRA: no se encontró el último informe monetario; " + " | ".join(errors))


def _existing_official_private_ratios(root: Path) -> list[dict[str, str]]:
    target = root / "data" / "processed" / "credit.csv"
    if not target.exists():
        return []
    with target.open(encoding="utf-8", newline="") as handle:
        return [
            row for row in csv.DictReader(handle)
            if row["series_id"] in {
                "bcra_private_nonfinancial_credit_gdp_ratio",
                "bcra_private_nonfinancial_credit_ars_gdp_ratio",
                "bcra_private_nonfinancial_credit_fx_ars_gdp_ratio",
            }
            and row["source_id"] == "bcra_monthly_monetary_indicators"
        ]


def _load_series(path: Path, series_id: str) -> tuple[dict[str, Decimal], str]:
    if not path.exists():
        raise PipelineError(f"crédito BCRA: falta el insumo {path}")
    content = path.read_bytes()
    with path.open(encoding="utf-8", newline="") as handle:
        values = {
            row["period"]: Decimal(row["value"])
            for row in csv.DictReader(handle)
            if row["series_id"] == series_id
        }
    if not values:
        raise PipelineError(f"crédito BCRA: falta la serie {series_id} en {path}")
    return values, hashlib.sha256(content).hexdigest()


def _quarter_number(period: str) -> int:
    year, quarter = period.split("-Q")
    return int(year) * 4 + int(quarter) - 1


def _gdp_denominators(gdp: dict[str, Decimal]) -> dict[int, Decimal]:
    """Compatibilidad para otros pipelines que aún usan promedios de 4 trimestres."""
    ordered = sorted((_quarter_number(period), value) for period, value in gdp.items())
    values = dict(ordered)
    denominators: dict[int, Decimal] = {}
    for quarter, _ in ordered:
        window = [values.get(quarter - offset) for offset in range(4)]
        if all(value is not None for value in window):
            denominators[quarter] = sum((value for value in window if value is not None), Decimal()) / Decimal(4)
    return denominators


def _month_number(period: str) -> int:
    year, month = map(int, period.split("-"))
    return year * 12 + month - 1


def _monthly_gdp_denominators(
    gdp: dict[str, Decimal], anchors: dict[str, Decimal] | None = None,
) -> dict[int, Decimal]:
    """Interpola el PIB nominal anualizado a frecuencia mensual.

    El valor trimestral corriente del INDEC ya está expresado a tasa anual y se
    ubica en el mes central del trimestre. Los cocientes oficiales del BCRA se
    usan como anclas adicionales de su PIB mensual estimado (promedio móvil de
    tres meses), evitando extrapolar con un trimestre nominal rezagado.
    """
    points = {
        int(year) * 12 + (int(quarter) - 1) * 3 + 1: value
        for period, value in gdp.items()
        for year, quarter in [period.split("-Q")]
    }
    for period, value in (anchors or {}).items():
        points[_month_number(period)] = value
    ordered = sorted(points)
    denominators: dict[int, Decimal] = {}
    for left, right in zip(ordered, ordered[1:]):
        start, end = points[left], points[right]
        width = Decimal(right - left)
        for month in range(left, right):
            denominators[month] = start + (end - start) * Decimal(month - left) / width
    if ordered:
        denominators[ordered[-1]] = points[ordered[-1]]
    return denominators


def calculate_derived(
    records: list[dict[str, str]], inflation_path: Path, gdp_path: Path,
    official_ratios: list[dict[str, str]] | None = None,
) -> list[dict[str, str]]:
    cpi, inflation_hash = _load_series(inflation_path, "indec_ipc_general_index")
    gdp, gdp_hash = _load_series(gdp_path, "indec_gdp_current_quarterly")
    if REAL_BASE_PERIOD not in cpi:
        raise PipelineError(f"crédito BCRA: IPC base {REAL_BASE_PERIOD} ausente")

    levels = {
        (row["series_id"], row["period"]): Decimal(row["value"])
        for row in records
        if row["series_id"] in LEVEL_SERIES
    }
    bases = {
        series_id: levels.get((series_id, REAL_BASE_PERIOD))
        for series_id in LEVEL_SERIES
    }
    missing_bases = sorted(series_id for series_id, value in bases.items() if value is None)
    if missing_bases:
        raise PipelineError(f"crédito BCRA: faltan bases {REAL_BASE_PERIOD}: {', '.join(missing_bases)}")

    total_levels = {
        row["period"]: Decimal(row["value"])
        for row in records if row["series_id"] == "bcra_private_nonfinancial_credit"
    }
    ratio_anchors = {
        row["period"]: total_levels[row["period"]] / (Decimal(row["value"]) / Decimal(100))
        for row in (official_ratios or [])
        if row["series_id"] == "bcra_private_nonfinancial_credit_gdp_ratio"
        and row["period"] in total_levels and Decimal(row["value"]) > 0
    }
    denominators = _monthly_gdp_denominators(gdp, ratio_anchors)
    if not denominators:
        raise PipelineError("crédito BCRA: no se pudo interpolar el PIB nominal mensual")
    available_months = sorted(denominators)
    derived: list[dict[str, str]] = []
    for row in records:
        series_id, period = row["series_id"], row["period"]
        if series_id not in LEVEL_SERIES:
            continue
        current = Decimal(row["value"])
        if period in cpi and period >= min(cpi):
            base = bases[series_id]
            assert base is not None
            real_index = (current / cpi[period]) / (base / cpi[REAL_BASE_PERIOD]) * Decimal(100)
            derived.append(row | {
                "series_id": f"{series_id}_real_index",
                "value": format(real_index.quantize(Decimal("0.000001")), "f"),
                "unit": "index_dec_2019_100",
                "status": "calculated",
                "source_id": "datarg_bcra_credit_deflated_by_indec_cpi",
                "source_sha256": hashlib.sha256(
                    f"{row['source_sha256']}|{inflation_hash}".encode()
                ).hexdigest(),
            })

        period_month = _month_number(period)
        usable = [month for month in available_months if month <= period_month]
        if usable:
            denominator = denominators[usable[-1]]
            ratio = current / denominator * Decimal(100)
            derived.append(row | {
                "series_id": f"{series_id}_gdp_ratio",
                "value": format(ratio.quantize(Decimal("0.000001")), "f"),
                "unit": "percent_gdp",
                "status": "calculated",
                "source_id": "datarg_bcra_credit_over_monthly_nominal_gdp",
                "source_sha256": hashlib.sha256(
                    f"{row['source_sha256']}|{gdp_hash}".encode()
                ).hexdigest(),
            })
    return derived


def _promote(records: list[dict[str, str]], root: Path, run_id: str) -> dict[str, object]:
    records.sort(key=lambda row: (row["series_id"], row["period"]))
    keys = [(row["series_id"], row["period"]) for row in records]
    if len(keys) != len(set(keys)):
        raise PipelineError("crédito BCRA: claves duplicadas")
    private = [row for row in records if row["series_id"] == "bcra_private_nonfinancial_credit"]
    if not private or private[0]["period"] != "1999-12":
        raise PipelineError("crédito BCRA: cobertura privada inesperada")
    target_dir = root / "data" / "processed"
    log_dir = root / "data" / "logs" / "credit"
    target_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / "credit.csv"
    old: dict[tuple[str, str], str] = {}
    legacy_private_ratio_keys: set[tuple[str, str]] = set()
    if target.exists():
        with target.open(encoding="utf-8", newline="") as handle:
            old_rows = list(csv.DictReader(handle))
        old = {(row["series_id"], row["period"]): row["value"] for row in old_rows}
        legacy_private_ratio_keys = {
            (row["series_id"], row["period"]) for row in old_rows
            if row["series_id"] == "bcra_private_nonfinancial_credit_gdp_ratio"
            and row["source_id"] != "bcra_monthly_monetary_indicators"
        }
    new = {(row["series_id"], row["period"]): row["value"] for row in records}
    report = {
        "run_id": run_id, "rows": len(records), "series": len({row["series_id"] for row in records}),
        "min_period": min(row["period"] for row in records), "max_period": max(row["period"] for row in records),
        "created": len(new.keys() - old.keys()), "deleted": len(old.keys() - new.keys()),
        "modified": sum(old[key] != new[key] for key in old.keys() & new.keys()),
    }
    unexpected_deleted = old.keys() - new.keys() - legacy_private_ratio_keys
    if unexpected_deleted:
        raise PipelineError("crédito BCRA: la fuente eliminó observaciones existentes")
    fd, temporary = tempfile.mkstemp(prefix="credit-", suffix=".csv", dir=target_dir)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
            writer.writeheader()
            writer.writerows(records)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)
    (log_dir / f"{run_id}.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return report


def run(
    root: Path,
    private_file: Path | None = None,
    private_currency_file: Path | None = None,
    public_file: Path | None = None,
    securities_file: Path | None = None,
    indicators_file: Path | None = None,
) -> dict[str, object]:
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    raw = root / "data" / "raw"
    private = acquire("bcra_private_nonfinancial_credit_xls", PRIVATE_URL, raw, private_file)
    private_currency = acquire(
        "bcra_private_credit_by_currency_xlsm", PRIVATE_CURRENCY_URL, raw, private_currency_file,
    )
    public = acquire("bcra_public_sector_loans_xls", PUBLIC_URL, raw, public_file)
    securities = acquire("bcra_public_securities_xls", SECURITIES_URL, raw, securities_file)
    indicators = _acquire_monthly_indicators(raw, indicators_file)
    records = extract(private, private_currency, public, securities)
    current_official_ratios = _official_private_credit_gdp(indicators)
    retained_official_ratios = _existing_official_private_ratios(root)
    all_official_ratios = retained_official_ratios + current_official_ratios
    records.extend(calculate_derived(
        records,
        root / "data" / "processed" / "inflation.csv",
        root / "data" / "processed" / "gdp.csv",
        all_official_ratios,
    ))
    official_ratios = {
        (row["series_id"], row["period"]): row
        for row in all_official_ratios
    }
    # Los puntos publicados por el BCRA reemplazan al cálculo de DatArg en las
    # fechas disponibles y actúan como control metodológico de la serie larga.
    merged = {(row["series_id"], row["period"]): row for row in records}
    merged.update(official_ratios)
    return _promote(list(merged.values()), root, run_id)
