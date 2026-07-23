from __future__ import annotations

import csv
import json
import os
import tempfile
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .inflation import OUTPUT_COLUMNS, Artifact, PipelineError, acquire

SOURCE_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/ITCRMSerie.xlsx"
SOURCE_ID = "bcra_itcrm_historical_xlsx"
SHEET_NAME = "ITCRM y bilaterales"
BASE_LABEL = "Índices con base 17-12-15=100"
FIRST_PERIOD = "1997-01-01"

# El BCRA destaca estos cuatro bilaterales en la página pública del indicador.
SERIES = {
    "ITCRM": "bcra_itcrm",
    "ITCRB Brasil": "bcra_itcrb_brazil",
    "ITCRB Estados Unidos": "bcra_itcrb_united_states",
    "ITCRB China": "bcra_itcrb_china",
    "ITCRB Zona Euro": "bcra_itcrb_euro_area",
}


def _text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _period(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return date.fromisoformat(str(value).strip()[:10]).isoformat()
    except ValueError as exc:
        raise PipelineError(f"ITCRM BCRA: período inválido: {value!r}") from exc


def _value(value: object, period: str, header: str) -> Decimal:
    try:
        number = Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise PipelineError(f"ITCRM BCRA: valor inválido en {header}/{period}") from exc
    if not Decimal("1") <= number <= Decimal("1000"):
        raise PipelineError(f"ITCRM BCRA: valor fuera de rango en {header}/{period}")
    return number


def extract(artifact: Artifact) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(artifact.path, read_only=True, data_only=True)
    except Exception as exc:
        raise PipelineError(f"ITCRM BCRA: no se pudo abrir el XLSX: {exc}") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise PipelineError(f"ITCRM BCRA: no existe la hoja {SHEET_NAME!r}")

    sheet = workbook[SHEET_NAME]
    if _text(sheet.cell(1, 1).value) != BASE_LABEL:
        raise PipelineError("ITCRM BCRA: cambió la base o el encabezado del libro")

    headers = {_text(cell.value): cell.column for cell in sheet[2] if _text(cell.value)}
    if _text(sheet.cell(2, 1).value) != "Período":
        raise PipelineError("ITCRM BCRA: falta la columna Período")
    missing = set(SERIES) - set(headers)
    if missing:
        raise PipelineError(f"ITCRM BCRA: faltan columnas esperadas: {sorted(missing)}")

    records: list[dict[str, str]] = []
    seen_periods: set[str] = set()
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            break
        period = _period(row[0])
        if period in seen_periods:
            raise PipelineError(f"ITCRM BCRA: período duplicado: {period}")
        seen_periods.add(period)
        for header, series_id in SERIES.items():
            value = _value(row[headers[header] - 1], period, header)
            records.append({
                "series_id": series_id,
                "period": period,
                "frequency": "daily",
                "value": format(value.quantize(Decimal("0.000001")), "f"),
                "unit": "index_dec_17_2015_100",
                "status": "official_provisional",
                "source_id": SOURCE_ID,
                "source_url": artifact.url,
                "source_sha256": artifact.sha256,
                "retrieved_at": artifact.retrieved_at,
            })
    workbook.close()
    if not records:
        raise PipelineError("ITCRM BCRA: la fuente no contiene observaciones")
    return records


def promote(records: list[dict[str, str]], root: Path, run_id: str) -> dict[str, object]:
    records.sort(key=lambda row: (row["series_id"], row["period"]))
    keys = [(row["series_id"], row["period"]) for row in records]
    if len(keys) != len(set(keys)):
        raise PipelineError("ITCRM BCRA: claves duplicadas")

    periods_by_series: dict[str, set[str]] = {}
    for row in records:
        periods_by_series.setdefault(row["series_id"], set()).add(row["period"])
    if set(periods_by_series) != set(SERIES.values()):
        raise PipelineError("ITCRM BCRA: panel de series incompleto")
    reference = periods_by_series["bcra_itcrm"]
    if any(periods != reference for periods in periods_by_series.values()):
        raise PipelineError("ITCRM BCRA: los bilaterales no comparten la misma cobertura")
    periods = sorted(reference)
    if periods[0] != FIRST_PERIOD or len(periods) < 10_000:
        raise PipelineError("ITCRM BCRA: cobertura histórica inesperada")

    target_dir = root / "data" / "processed"
    log_dir = root / "data" / "logs" / "real_exchange_rate"
    target_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / "real_exchange_rate.csv"

    old_keys: set[tuple[str, str]] = set()
    old_values: dict[tuple[str, str], str] = {}
    if target.exists():
        with target.open(encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                key = (row["series_id"], row["period"])
                old_keys.add(key)
                old_values[key] = row["value"]
    new_values = {(row["series_id"], row["period"]): row["value"] for row in records}
    new_keys = set(new_values)
    report = {
        "run_id": run_id,
        "rows": len(records),
        "series": len(periods_by_series),
        "min_period": periods[0],
        "max_period": periods[-1],
        "created": len(new_keys - old_keys),
        "deleted": len(old_keys - new_keys),
        "modified": sum(old_values[key] != new_values[key] for key in old_keys & new_keys),
    }
    if old_keys and report["deleted"]:
        raise PipelineError("ITCRM BCRA: la fuente eliminó observaciones existentes")

    fd, temporary = tempfile.mkstemp(prefix="real-exchange-rate-", suffix=".csv", dir=target_dir)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
            writer.writeheader()
            writer.writerows(records)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)

    (log_dir / f"{run_id}.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return report


def run(root: Path, source_file: Path | None = None) -> dict[str, object]:
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    artifact = acquire(SOURCE_ID, SOURCE_URL, root / "data" / "raw", source_file)
    return promote(extract(artifact), root, run_id)
